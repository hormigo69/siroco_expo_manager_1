import os
import glob
import shutil
import re
import pandas as pd
from PIL import Image, ExifTags
import subprocess
import json
import concurrent.futures
from tqdm import tqdm
import time
import platform
import sys
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from flask import Flask
from multiprocessing import Manager

app = Flask(__name__)

#activar el entorno virtual
#source venv/bin/activate   


class FileManager:
    @staticmethod
    #Asegura que la carpeta exista, si no, la crea
    def ensure_directory(path):
        if not os.path.exists(path):
            os.makedirs(path)

    @staticmethod
    def unzip_output(zip_path, output_path):
        import zipfile
        import os

        obras_path = os.path.join(output_path, 'Obras')
        os.makedirs(obras_path, exist_ok=True)

        # Buscar todos los archivos zip en la carpeta
        zip_files = glob.glob(os.path.join(output_path, "*.zip"))
        
        for zip_path in zip_files:
            print(f"Procesando archivo zip: {zip_path}")
            
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                for file_info in zip_ref.infolist():
                    if file_info.filename.startswith('__MACOSX') or file_info.filename.startswith('.'):
                        continue  # Saltar archivos específicos de macOS y archivos ocultos
                    
                    # Mantener la estructura de carpetas interna
                    target_path = os.path.join(obras_path, file_info.filename)
                    
                    # Saltar directorios vacos
                    if file_info.filename.endswith('/'):
                        continue

                    # Extraer el archivo
                    os.makedirs(os.path.dirname(target_path), exist_ok=True)
                    
                    try:
                        with zip_ref.open(file_info) as source, open(target_path, "wb") as target:
                            shutil.copyfileobj(source, target)
                    except Exception as e:
                        print(f"Error al extraer {file_info.filename}: {str(e)}")

        print(f"Contenido descomprimido en: {obras_path}")


    @staticmethod
    #Copia los ficheros de la carpeta origen a la carpeta destino, renombrando los ficheros según el formato de la pantalla
    def process_files(carpeta_origen, carpeta_destino, id_expo):
        FileManager.ensure_directory(carpeta_destino)
        varios_folder = os.path.join(carpeta_destino, 'varios')
        FileManager.ensure_directory(varios_folder)
        
        for root, _, files in os.walk(carpeta_origen):
            for file_name in files:
                if file_name == '.DS_Store':
                    continue
                
                file_path = os.path.join(root, file_name)
                parent_folder = os.path.basename(os.path.dirname(file_path))
                
                if parent_folder.startswith('PANTALLA'):
                    pantalla_num = parent_folder.split()[1].zfill(3)
                    new_file_name = f"{id_expo}_P{pantalla_num}_{file_name}"
                    new_file_path = os.path.join(carpeta_destino, new_file_name)
                else:
                    new_file_path = os.path.join(varios_folder, file_name)
                
                try:
                    shutil.copy2(file_path, new_file_path)
                except Exception as e:
                    print(f"Error al copiar {file_path}: {str(e)}")

    @staticmethod
    #Mueve los ficheros que no coinciden con el formato de la pantalla a la carpeta varios
    def move_non_matching_files(carpeta_destino, varios_folder, expo_id):
        patron = re.compile(f"^{expo_id}_P\d{{3}}_")
        for file in os.listdir(carpeta_destino):
            file_path = os.path.join(carpeta_destino, file)
            if os.path.isfile(file_path) and not patron.match(file):
                destino = os.path.join(varios_folder, file)
                try:
                    shutil.move(file_path, destino)
                except Exception as e:
                    print(f"Error al mover {file}: {str(e)}")

    @staticmethod
    def is_video(file_path):
        video_extensions = ['.mp4', '.avi', '.mov', '.mkv', '.flv', '.wmv']
        return any(file_path.lower().endswith(ext) for ext in video_extensions)

class ImageInfo:
    @staticmethod
    def get_image_info(file_path):
        try:
            with Image.open(file_path) as img:
                width, height = img.size
                
                # Manejo especial de DPI para PNG
                if img.format == 'PNG':
                    # Buscar la información de DPI en los metadatos específicos de PNG
                    dpi = None
                    if 'dpi' in img.info:
                        dpi = img.info['dpi']
                    elif 'pHYs' in img.info:
                        # pHYs contiene pixels por metro, convertir a DPI
                        x_density = img.info['pHYs'][0]
                        y_density = img.info['pHYs'][1]
                        if x_density > 0 and y_density > 0:
                            # Convertir de pixels por metro a DPI (1 metro = 39.3701 pulgadas)
                            dpi = (round(x_density / 39.3701), round(y_density / 39.3701))
                else:
                    dpi = img.info.get('dpi', (None, None))

                # Si no se encontró DPI, usar valor por defecto de 72
                if not dpi or None in dpi:
                    dpi = (72, 72)
                
                img_format = img.format
                file_size = os.path.getsize(file_path)
                file_size_mb = file_size / (1024 * 1024)
                
                orientation = 'H' if width > height else 'V'
                
                return {
                    'ANCHO': width,
                    'ALTO': height,
                    'RESOLUCION_X': dpi[0],
                    'RESOLUCION_Y': dpi[1],
                    'ORIENTACION': orientation,
                    'FORMATO': img_format,
                    'TAMAÑO_MB': round(file_size_mb, 2)
                }
        except Exception as e:
            print(f"Error procesando {file_path}: {e}")
            return None
                
    @staticmethod
    def get_images_info_from_directory(directory_path):
        '''
        Recorre solo el directorio raíz especificado y crea un DataFrame con las características de las imágenes.
        '''
        images_info = []
        
        for file_name in os.listdir(directory_path):
            file_path = os.path.join(directory_path, file_name)
            
            # Ignorar subdirectorios
            if os.path.isdir(file_path):
                continue
            
            # Intentar abrir el archivo como imagen
            try:
                with Image.open(file_path):
                    pass  # Si no lanza una excepción, es una imagen
            except IOError:
                # No es una imagen o está corrupta
                print(f"Archivo no es una imagen o está corrupto: {file_path}")
                continue
            
            # Obtener información de la imagen
            info = ImageInfo.get_image_info(file_path)
            if info:
                info['NOMBRE_ARCHIVO'] = file_name
                
                # Extraer el número de pantalla
                match = re.search(r'_P(\d{3})_', file_name)
                if match:
                    info['NUMERO_PANTALLA'] = int(match.group(1))
                else:
                    info['NUMERO_PANTALLA'] = None
                
                images_info.append(info)
    
        return pd.DataFrame(images_info)

class VideoInfo:
    @staticmethod
    def get_video_info(file_path):
        '''
        Obtiene la información de un archivo de video usando ffprobe.
        '''
        try:
            result = subprocess.run(['ffprobe', '-v', 'quiet', '-print_format', 'json', '-show_format', '-show_streams', file_path], 
                                    stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            probe = json.loads(result.stdout)
            
            video_stream = next((stream for stream in probe['streams'] if stream['codec_type'] == 'video'), None)
            
            if video_stream:
                width = int(video_stream['width'])
                height = int(video_stream['height'])
                orientation = 'H' if width > height else 'V'
                # Redondear los FPS a un decimal
                frame_rate = round(float(eval(video_stream['r_frame_rate'])), 1)
                duration = float(probe['format']['duration'])
                duration_seconds = int(duration)
                bit_rate = int(probe['format']['bit_rate'])
                video_codec = video_stream['codec_name']
                file_extension = os.path.splitext(file_path)[1][1:].lower()

                file_size = os.path.getsize(file_path)
                file_size_mb = file_size / (1024 * 1024)  # Convertir a MB

                return {
                    'ANCHO': width,
                    'ALTO': height,
                    'ORIENTACION': orientation,
                    'DURACION_SEG': duration_seconds,
                    'FPS': frame_rate,  # Ahora frame_rate ya está redondeado
                    'FORMATO': file_extension,
                    'TAMAÑO_MB': round(file_size_mb, 2),
                    'TASA_BITS': bit_rate,
                    'CÓDEC_VIDEO': video_codec,
                    'NOMBRE_ARCHIVO': os.path.basename(file_path)
                }
        except Exception as e:
            print(f"Error procesando el video {file_path}: {e}")
            return None

def get_file_info(file_path):
    if FileManager.is_video(file_path):
        info = VideoInfo.get_video_info(file_path)
        if info:
            info['TIPO'] = 'Video'
        return info
    else:
        info = ImageInfo.get_image_info(file_path)
        if info:
            info['TIPO'] = 'Imagen'
        return info

def get_files_info_from_directory(directory_path):
    files_info = []
    
    for file_name in os.listdir(directory_path):
        file_path = os.path.join(directory_path, file_name)
        
        # Ignorar subdirectorios
        if os.path.isdir(file_path):
            continue
        
        # Obtener información del archivo
        info = get_file_info(file_path)
        if info:
            info['NOMBRE_ARCHIVO'] = file_name
            
            # Extraer el número de pantalla
            match = re.search(r'_P(\d{3})_', file_name)
            if match:
                info['NUMERO_PANTALLA'] = int(match.group(1))
            else:
                info['NUMERO_PANTALLA'] = None
            
            files_info.append(info)

    df = pd.DataFrame(files_info)
    if df.empty:
        print("No se pudo procesar ningún archivo correctamente.")
        return pd.DataFrame(columns=['NOMBRE_ARCHIVO', 'NUMERO_PANTALLA'])
    return df

class ExpoProcessor:
    def __init__(self, expo_id):
        self.expo_id = expo_id
        self.output_folder = expo_id
        self.full_output_path = os.path.join(os.getcwd(), 'expos', self.output_folder)
        self.carpeta_origen = os.path.join(self.full_output_path, 'Obras')
        self.carpeta_destino = os.path.join(self.full_output_path, 'ficheros_salida')
        self.varios_folder = os.path.join(self.carpeta_destino, 'varios')

    def setup_directories(self):
        FileManager.ensure_directory(os.path.join(self.full_output_path, 'Obras'))
        FileManager.ensure_directory(os.path.join(self.full_output_path, 'ficheros_salida'))
        FileManager.ensure_directory(self.varios_folder)

    def unzip_files(self):
        zip_files = glob.glob(os.path.join(self.full_output_path, "*.zip"))
        if zip_files:
            zip_path = zip_files[0]
            print(f"Archivo zip encontrado: {zip_path}")
            FileManager.unzip_output(zip_path, self.full_output_path)
        else:
            print("No se encontraron archivos .zip en la carpeta.")

    def process_and_move_files(self):
        FileManager.process_files(self.carpeta_origen, self.carpeta_destino, self.expo_id)
        FileManager.move_non_matching_files(self.carpeta_destino, self.varios_folder, self.expo_id)
        shutil.rmtree(self.carpeta_origen)
        print("Proceso de movimiento de archivos completado.")

    def generate_summary(self):
        # Importar las librerías necesarias
        from openpyxl import load_workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo

        df = get_files_info_from_directory(self.carpeta_destino)
        
        if not df.empty:
            if 'NUMERO_PANTALLA' in df.columns:
                df = df.sort_values('NUMERO_PANTALLA')
            
            excel_path = os.path.join(self.full_output_path, 'Resumen obras.xlsx')
            
            # Guardar primero con pandas
            df.to_excel(excel_path, index=False)
            
            # Abrir el archivo con openpyxl para añadir el formato de tabla
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Obtener el rango de la tabla
            tabla_ref = f"A1:{chr(64 + len(df.columns))}{len(df.index) + 1}"
            
            # Crear la tabla con estilo
            tabla = Table(displayName="TablaObras", ref=tabla_ref)
            estilo = TableStyleInfo(
                name="TableStyleMedium2", 
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tabla.tableStyleInfo = estilo
            
            # Añadir la tabla a la hoja
            ws.add_table(tabla)
            
            # Ajustar el ancho de las columnas
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Guardar los cambios
            wb.save(excel_path)
            print("Proceso de resumen de archivos completado con formato de tabla.")
        else:
            print("No se encontraron archivos para procesar.")

class Recodificador:
    def __init__(self, expo_id):
        self.expo_id = expo_id
        self.full_output_path = os.path.join(os.getcwd(), 'expos', self.expo_id)
        self.carpeta_origen = os.path.join(self.full_output_path, 'ficheros_salida')
        self.carpeta_destino = os.path.join(self.full_output_path, 'procesados')
        self.max_workers = max(1, min(os.cpu_count() - 1, 4))
        
        # Crear un archivo de progreso con lock
        self.progress_file = os.path.join(self.full_output_path, 'progress.json')
        if os.path.exists(self.progress_file):
            os.remove(self.progress_file)
        with open(self.progress_file, 'w') as f:
            json.dump({}, f)

    def process_file(self, file_path):
        try:
            file_name = os.path.basename(file_path)
            print(f"Iniciando procesamiento de {file_name}")
            
            if file_name.lower().endswith('.mp4'):
                print(f"Procesando video: {file_name}")
                archivo_salida = os.path.join(self.carpeta_destino, file_name)
                
                cmd = [
                    'ffmpeg',
                    '-i', file_path,
                    '-vf', 'fps=30,scale=1920:1080:flags=bicubic',
                    '-c:v', 'libx265',
                    '-preset', 'medium',
                    '-crf', '23',
                    '-c:a', 'aac',
                    '-b:v', '45M',
                    '-maxrate', '60M',
                    '-bufsize', '60M',
                    '-movflags', '+faststart',
                    '-progress', '-',  # Enviar progreso a stdout
                    '-y',
                    archivo_salida
                ]
                
                print(f"Ejecutando comando: {' '.join(cmd)}")
                process = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    universal_newlines=True,
                    bufsize=1  # Line buffered
                )
                
                start_time = time.time()
                last_progress_update = 0
                
                while True:
                    # Leer la salida de ffmpeg
                    output = process.stdout.readline()
                    if output == '' and process.poll() is not None:
                        break
                    
                    # Imprimir la salida para debug
                    if output:
                        print(f"[{file_name}] {output.strip()}")
                        
                        # Calcular progreso basado en el tiempo transcurrido
                        current_time = time.time()
                        elapsed_time = current_time - start_time
                        
                        # Actualizar progreso cada segundo
                        if elapsed_time - last_progress_update >= 1:
                            progress = min(95, (elapsed_time / 300) * 100)  # 300 segundos estimados
                            self.save_progress(file_name, progress)
                            last_progress_update = elapsed_time
                            print(f"Progreso de {file_name}: {progress:.2f}%")
                    
                    # Leer también stderr para evitar bloqueos
                    error = process.stderr.readline()
                    if error:
                        print(f"[{file_name}] Error: {error.strip()}")
                
                # Obtener el código de retorno
                return_code = process.poll()
                
                if return_code == 0:
                    print(f"Video {file_name} procesado exitosamente")
                    self.save_progress(file_name, 100)
                    return True
                else:
                    print(f"Error procesando video {file_name}. Código de retorno: {return_code}")
                    error_output = process.stderr.read()
                    print(f"Error detallado: {error_output}")
                    self.save_progress(file_name, 0)
                    return False
                    
            else:
                print(f"Procesando imagen: {file_name}")
                self.procesar_imagen(file_path)
                self.save_progress(file_name, 100)
                return True
                
        except Exception as e:
            print(f"Error procesando {file_path}: {str(e)}")
            self.save_progress(file_name, 0)
            return False

    def save_progress(self, filename, progress):
        """Guarda el progreso de forma segura"""
        max_retries = 3
        for _ in range(max_retries):
            try:
                with open(self.progress_file, 'r') as f:
                    progress_dict = json.load(f)
                progress_dict[filename] = progress
                with open(self.progress_file, 'w') as f:
                    json.dump(progress_dict, f)
                return True
            except Exception as e:
                print(f"Error guardando progreso: {e}")
                time.sleep(0.1)
        return False

    def procesar_video(self, file_path):
        try:
            nombre_archivo = os.path.basename(file_path)
            archivo_salida = os.path.join(self.carpeta_destino, nombre_archivo)
            print(f"Procesando video {nombre_archivo} -> {archivo_salida}")
            
            # Primero obtener la duración del video
            probe_cmd = [
                'ffprobe', 
                '-v', 'error',
                '-show_entries', 'format=duration',
                '-of', 'default=noprint_wrappers=1:nokey=1',
                file_path
            ]
            
            duration = float(subprocess.check_output(probe_cmd).decode().strip())
            print(f"Duración del video: {duration} segundos")
            
            # Configurar el comando ffmpeg
            cmd = [
                'ffmpeg',
                '-i', file_path,
                '-vf', 'fps=30,scale=1920:1080:flags=bicubic',
                '-c:v', 'libx265',
                '-preset', 'medium',
                '-crf', '23',
                '-c:a', 'aac',
                '-b:v', '45M',
                '-maxrate', '60M',
                '-bufsize', '60M',
                '-movflags', '+faststart',
                '-progress', 'pipe:1',  # Enviar progreso a stdout
                '-y',
                archivo_salida
            ]

            print(f"Ejecutando comando: {' '.join(cmd)}")
            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                universal_newlines=True
            )
            
            # Leer la salida línea por línea para monitorear el progreso
            while True:
                line = process.stdout.readline()
                if not line:
                    break
                
                # Buscar la línea que contiene el tiempo actual
                if 'out_time_ms=' in line:
                    current_time = float(line.split('=')[1]) / 1000000  # Convertir microsegundos a segundos
                    progress = min(100, (current_time / duration) * 100)
                    print(f"Progreso de {nombre_archivo}: {progress:.2f}%")
            
            # Esperar a que termine el proceso
            process.wait()
            
            if process.returncode != 0:
                stderr = process.stderr.read()
                print(f"Error en ffmpeg: {stderr}")
                raise subprocess.CalledProcessError(process.returncode, cmd, stderr)
            
            print(f"Video procesado exitosamente: {nombre_archivo}")
            return True
            
        except Exception as e:
            print(f"Error procesando video {file_path}: {str(e)}")
            raise

    def crear_carpeta_procesados(self):
        if not os.path.exists(self.carpeta_destino):
            os.makedirs(self.carpeta_destino)
            print(f"Carpeta '{self.carpeta_destino}' creada.")
        else:
            # Si la carpeta existe, limpiarla
            for archivo in os.listdir(self.carpeta_destino):
                ruta_archivo = os.path.join(self.carpeta_destino, archivo)
                if os.path.isfile(ruta_archivo):
                    os.remove(ruta_archivo)
            print(f"Carpeta '{self.carpeta_destino}' limpiada para nuevo procesamiento.")
        return True

    def procesar_archivos(self):
        archivos = os.listdir(self.carpeta_origen)
        total_archivos = len(archivos)

        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = []
            for archivo in archivos:
                ruta_archivo = os.path.join(self.carpeta_origen, archivo)
                if archivo.lower().endswith(('.png', '.jpg', '.jpeg')):
                    futures.append(executor.submit(self.procesar_imagen, ruta_archivo))
                elif archivo.lower().endswith('.mp4'):
                    futures.append(executor.submit(self.procesar_video, ruta_archivo))

            with tqdm(total=total_archivos, desc="Progreso global") as pbar:
                for future in concurrent.futures.as_completed(futures):
                    pbar.update(1)

    def procesar_imagen(self, ruta_archivo):
        nombre_archivo = os.path.basename(ruta_archivo)
        with tqdm(total=1, desc=f"Procesando imagen: {nombre_archivo}", leave=False) as pbar:
            try:
                with Image.open(ruta_archivo) as img:
                    ancho, alto = img.size
                    formato = img.format
                    modo = img.mode
                    tamano = os.path.getsize(ruta_archivo)
                    tamano_kb = tamano / 1024

                    necesita_procesamiento = False
                    razones = []

                    # Verificar resolución (debe ser exactamente 2160x3840 o 3840x2160)
                    if (ancho, alto) not in [(2160, 3840), (3840, 2160)]:
                        necesita_procesamiento = True
                        razones.append(f"Resolución incorrecta: {ancho}x{alto}")

                    # Verificar formato y modo
                    if formato not in ['JPEG', 'JPG', 'PNG']:
                        necesita_procesamiento = True
                        razones.append(f"Formato incorrecto: {formato}")
                    if modo != 'RGB':
                        necesita_procesamiento = True
                        razones.append(f"Modo de color incorrecto: {modo}")

                    # Verificar tamaño (más de 10MB sería excesivo para estas resoluciones)
                    if tamano_kb > 10 * 1024:  # más de 10MB
                        necesita_procesamiento = True
                        razones.append(f"Tamaño excesivo: {tamano_kb:.2f}KB")

                    if necesita_procesamiento:
                        print(f"Procesando {nombre_archivo} por: {', '.join(razones)}")
                        img_procesada = img.convert('RGB')
                        
                        # Ajustar resolución si es necesario
                        if (ancho, alto) not in [(2160, 3840), (3840, 2160)]:
                            if ancho > alto:
                                nuevo_ancho, nuevo_alto = 3840, 2160
                            else:
                                nuevo_ancho, nuevo_alto = 2160, 3840
                            img_procesada = img_procesada.resize((nuevo_ancho, nuevo_alto), Image.Resampling.LANCZOS)

                        nuevo_nombre = os.path.splitext(nombre_archivo)[0] + '_procesado.jpg'
                        ruta_destino = os.path.join(self.carpeta_destino, nuevo_nombre)
                        img_procesada.save(ruta_destino, 'JPEG', 
                                         quality=85,  # Alta calidad pero no máxima
                                         optimize=True,  # Optimizar el encoding
                                         dpi=(72, 72),  # Establecer DPI a 72
                                         progressive=True)  # Hacer la imagen progresiva
                    else:
                        print(f"No se procesa {nombre_archivo}: cumple todos los criterios")
                        print(f"- Formato: {formato}")
                        print(f"- Modo: {modo}")
                        print(f"- Resolución: {ancho}x{alto}")
                        print(f"- Tamaño: {tamano_kb:.2f}KB")
                        nuevo_nombre = nombre_archivo
                        ruta_destino = os.path.join(self.carpeta_destino, nuevo_nombre)
                        shutil.copy2(ruta_archivo, ruta_destino)

                time.sleep(0.1)
                pbar.update(1)
                return nuevo_nombre

            except Exception as e:
                print(f"Error al procesar la imagen {nombre_archivo}: {str(e)}")
                return nombre_archivo
            
    def actualizar_excel(self):
        try:
            # Leer el Excel existente
            df = pd.read_excel(self.excel_path)
            
            # Obtener información actualizada de los archivos procesados
            archivos_procesados = get_files_info_from_directory(self.carpeta_destino)
            
            # Crear un diccionario para mapear nombres de archivo a sus nuevos datos
            nuevos_datos = {row['NOMBRE_ARCHIVO']: row for _, row in archivos_procesados.iterrows()}
            
            # Columnas para los datos recodificados
            columnas_recodificadas = ['ANCHO_RECODIFICADO', 'ALTO_RECODIFICADO', 'RESOLUCION_X_RECODIFICADO', 
                                      'RESOLUCION_Y_RECODIFICADO', 'ORIENTACION_RECODIFICADO', 'FORMATO_RECODIFICADO', 
                                      'TAMAÑO_MB_RECODIFICADO', 'DURACION_SEG_RECODIFICADO', 'FPS_RECODIFICADO', 
                                      'TASA_BITS_RECODIFICADO', 'CÓDEC_VIDEO_RECODIFICADO']
            
            # Añadir columnas para los datos recodificados
            for columna in columnas_recodificadas:
                if columna not in df.columns:
                    df[columna] = None
            
            # Actualizar el DataFrame con los nuevos datos
            for index, row in df.iterrows():
                nombre_archivo = row['NOMBRE_ARCHIVO']
                nombre_procesado = nombre_archivo.replace('.', '_procesado.')
                
                if nombre_procesado in nuevos_datos:
                    df.at[index, 'PROCESADO'] = 'Sí'
                    for columna in columnas_recodificadas:
                        columna_original = columna.replace('_RECODIFICADO', '')
                        if columna_original in nuevos_datos[nombre_procesado]:
                            df.at[index, columna] = nuevos_datos[nombre_procesado][columna_original]
                else:
                    df.at[index, 'PROCESADO'] = 'No'
            
            # Guardar primero con pandas
            df.to_excel(self.excel_path, index=False)
            
            # Abrir el archivo con openpyxl para añadir el formato de tabla
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Obtener el rango de la tabla
            tabla_ref = f"A1:{chr(64 + len(df.columns))}{len(df.index) + 1}"
            
            # Crear la tabla con estilo
            tabla = Table(displayName="TablaObras", ref=tabla_ref)
            estilo = TableStyleInfo(
                name="TableStyleMedium2", 
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tabla.tableStyleInfo = estilo
            
            # Añadir la tabla a la hoja
            ws.add_table(tabla)
            
            # Ajustar el ancho de las columnas
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Guardar los cambios
            wb.save(self.excel_path)
            print(f"Excel actualizado con formato de tabla: {self.excel_path}")
        
        except Exception as e:
            print(f"Error al actualizar el Excel: {str(e)}")

def main():
    expo_id = 'E995'
    recodificador = Recodificador(expo_id)
    carpeta_procesados = os.path.join(os.getcwd(), 'expos', expo_id, 'procesados')
    
    # Verificar si ya existen archivos procesados
    if os.path.exists(carpeta_procesados) and os.listdir(carpeta_procesados):
        while True:
            print("Se encontraron archivos procesados. ¿Quieres volver a procesar los archivos? (s/n): ", end='', flush=True)
            try:
                # Leemos la línea completa y eliminamos espacios y saltos de línea
                respuesta = input().strip()
                
                if respuesta == 's':
                    break
                elif respuesta == 'n':
                    print("Actualizando solo el Excel con los archivos existentes...")
                    recodificador.actualizar_excel()
                    return
                print("Por favor, introduce 's' o 'n'")
            except EOFError:
                continue


    # Si no hay archivos procesados o el usuario quiere reprocesar, continuar con el flujo normal
    processor = ExpoProcessor(expo_id)
    processor.setup_directories()
    processor.unzip_files()
    processor.process_and_move_files()
    processor.generate_summary()

    # Proceso de recodificación
    recodificador.crear_carpeta_procesados()
    recodificador.procesar_archivos()
    recodificador.actualizar_excel()

if __name__ == "__main__":

    # El resto del código main() se ejecutará si se llama directamente
    main()
    app.run(debug=True, port=5000)